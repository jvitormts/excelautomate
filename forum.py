
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def load_workbook_sheet(file_path, sheet_name=None):
    """
    Carrega uma planilha de um arquivo Excel.
    
    :param file_path: Caminho do arquivo Excel.
    :param sheet_name: Nome da aba a ser carregada (None para carregar a aba ativa).
    :return: A aba carregada do arquivo.
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    return workbook, sheet

def copy_data_columns(source_sheet, dest_sheet, source_col_start, dest_col_start, rows):
    """
    Copia dados entre colunas específicas de origem e destino, com validação de dados copiados.
    
    :param source_sheet: Aba de origem.
    :param dest_sheet: Aba de destino.
    :param source_col_start: Número da coluna inicial na planilha de origem.
    :param dest_col_start: Número da coluna inicial na planilha de destino.
    :param rows: Número de linhas a serem copiadas.
    """
    for i in range(1, rows + 1):
        source_cell_1 = source_sheet[f'{get_column_letter(source_col_start)}{i + 3}']
        dest_cell_1 = dest_sheet[f'{get_column_letter(dest_col_start)}{i + 2}']
        dest_cell_1.value = source_cell_1.value
        print(f'Copiado {source_cell_1.value} da célula {source_cell_1.coordinate} para {dest_cell_1.coordinate}')

        source_cell_2 = source_sheet[f'{get_column_letter(source_col_start + 1)}{i + 3}']
        dest_cell_2 = dest_sheet[f'{get_column_letter(dest_col_start + 1)}{i + 2}']
        dest_cell_2.value = source_cell_2.value
        print(f'Copiado {source_cell_2.value} da célula {source_cell_2.coordinate} para {dest_cell_2.coordinate}')

def copy_data_range(source_sheet, dest_sheet, source_start_cell, dest_start_cell, rows, cols):
    """
    Copia dados de uma faixa específica de células, com validação de dados copiados.
    
    :param source_sheet: Aba de origem.
    :param dest_sheet: Aba de destino.
    :param source_start_cell: Célula inicial na origem (ex: 'A1').
    :param dest_start_cell: Célula inicial no destino (ex: 'A1').
    :param rows: Número de linhas a serem copiadas.
    :param cols: Número de colunas a serem copiadas.
    """
    source_row_start = source_sheet[source_start_cell].row
    source_col_start = source_sheet[source_start_cell].column
    dest_row_start = dest_sheet[dest_start_cell].row
    dest_col_start = dest_sheet[dest_start_cell].column

    for row in range(rows):
        for col in range(cols):
            source_cell = source_sheet.cell(row=source_row_start + row, column=source_col_start + col)
            dest_cell = dest_sheet.cell(row=dest_row_start + row, column=dest_col_start + col)
            dest_cell.value = source_cell.value
            print(f'Copiado {source_cell.value} da célula {source_cell.coordinate} para {dest_cell.coordinate}')

def main(source_file, dest_file, rows=100):
    """
    Função principal que carrega as planilhas, copia os dados e salva a planilha de destino.
    
    :param source_file: Caminho do arquivo de origem.
    :param dest_file: Caminho do arquivo de destino.
    :param rows: Número de linhas a serem copiadas na primeira aba.
    """
    # Carregar aba "dados" de ambos os arquivos
    source_wb, source_sheet_dados = load_workbook_sheet(source_file, "dados")
    dest_wb, dest_sheet_dados = load_workbook_sheet(dest_file, "dados")

    # Copiar dados de "dados" (colunas H4 e H5 da origem para H3 e H4 no destino)
    copy_data_columns(source_sheet_dados, dest_sheet_dados, source_col_start=8, dest_col_start=8, rows=rows)

    # Carregar aba "Closed by application" e "por aplicação"
    source_sheet_application = source_wb["Closed by application"]
    dest_sheet_application = dest_wb["por aplicação"]

    # Copiar dados de "Closed by application" para "por aplicação" (12L x 2C)
    copy_data_range(source_sheet_application, dest_sheet_application, source_start_cell="A1", dest_start_cell="A1", rows=12, cols=2)

    # Salvar planilha de destino
    dest_wb.save(dest_file)
    print(f"Dados copiados de {source_file} para {dest_file} com sucesso!")

# Exemplo de uso
source_file_path = "caminho_para_planilha_origem.xlsx"
dest_file_path = "caminho_para_planilha_destino.xlsx"
main(source_file_path, dest_file_path, rows=100)
